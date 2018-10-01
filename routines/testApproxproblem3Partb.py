#!/usr/bin/env python3

import cmath, math
from openpyxl import Workbook
import approxDeriv, relError, absError, DmacEp

def EvaluateApprox():
    #initalize 
    i = 0
    h = 1
    
    aListS = []
    rListS= []
    aListC = []
    rListC = []
    difference  = []
    

    apListSin = []
    apListTaylor = []

    hList = []
    sinX = 0
    abC = 1
    relC = 1

    #We are creating a while loop that will reduce h by half every time. We will continue until abserror no longer improves

    while (sinX < 1):
        i += 1
        sinX = approxDeriv.approxDeriv(h, lambda x: math.sin(x), 0)
        taylorSin = approxDeriv.approxDeriv(h, lambda x: x - pow(x, 3)/6 + pow(x, 5)/120, 0)

        apListSin.append(sinX)
        apListTaylor.append(taylorSin)
        difference.append(abs(sinX-taylorSin))
        hList.append(h)
        h *= .5  
    
    #opening a workbook and adding the values of h, absError, relError into their own columns in xlsx file.
    wb = Workbook()
    sheet = wb.active
    r = 2
    for h in hList:
        sheet.cell(row= r, column=1).value = h
        r += 1
    r = 2
    for ap in apListSin:
        sheet.cell(row=r, column= 2).value = ap
        r += 1
    r = 2
    for a in apListTaylor:
        sheet.cell(row= r, column=3).value = a
        r += 1
    r = 2
    for a in difference:
        sheet.cell(row= r, column=4).value = a
        r += 1
    
    wb.save('TestApprox3b.xlsx')

    

EvaluateApprox()
