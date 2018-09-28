#!/usr/bin/env python3

import cmath
from openpyxl import Workbook
import approxDeriv, relError, absError, DmacEp

def EvaluateApprox():
    #initalize 
    i = 0
    h = 1
    aList = []
    rList = []
    apList = []
    hList = []
    ab = 1
    rel = 1

    #We are creating a while loop that will reduce h by half every time. We will continue until abserror no longer improves

    while (ab != 0):
        i += 1
        approxD = approxDeriv.approxDeriv(h, lambda x: x**3 + 2*x**2 - 3*x, 1)
        ab = absError.absError(approxD, 4)
        rel = relError.relError(approxD,4)
        aList.append(ab)
        rList.append(rel)
        hList.append(h)
        apList.append(approxD)
        print("{0:1E} {1:2E} {2:3E} {3:4E} {4:5E}".format(h, approxD, 4, ab, rel))
        h *= .5  
    
    #opening a workbook and adding the values of h, absError, relError into their own columns in xlsx file.
    wb = Workbook()
    sheet = wb.active
    r = 1
    for h in hList:
        sheet.cell(row= r, column=1).value = h
        r += 1
    r = 1
    for ap in apList:
        sheet.cell(row=r, column= 4).value = ap
        r += 1
    r = 1
    for a in aList:
        sheet.cell(row= r, column=2).value = a
        r += 1
    r = 1
    for re in rList:
        sheet.cell(row= r, column=3).value = re
        r += 1
    wb.save('TestApprox2.xlsx')

    

EvaluateApprox()
