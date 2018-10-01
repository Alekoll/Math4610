#!/usr/bin/env python3

import cmath
from openpyxl import Workbook
import approxDeriv, relError, absError, DmacEp

def EvaluateApprox():
    #initalize 
    i = 0
    h = 1
    
    aListS = []
    rListS= []
    aListC = []
    rListC = []

    apListS = []
    apListC = []

    hList = []
    abS = 1
    relS = 1
    abC = 1
    relC = 1

    #We are creating a while loop that will reduce h by half every time. We will continue until abserror no longer improves

    while (i < 40):
        i += 1
        sqrtX = approxDeriv.approxDeriv(h, lambda x: x**(1/2), 7)
        sqrtConjX = 1/((7 + h)**(1/2) + 7**(1/2))

        
        abS = absError.absError(sqrtX, cmath.sqrt(7)/(2*7))
        relS = relError.relError(sqrtX,cmath.sqrt(7)/(2*7))

        abC = absError.absError(sqrtConjX, cmath.sqrt(7)/(2*7))
        relC = relError.relError(sqrtConjX,cmath.sqrt(7)/(2*7))

        aListS.append(abS)
        rListS.append(relS)

        aListC.append(abC)
        rListC.append(relC)

        hList.append(h)
        apListS.append(sqrtX)
        apListC.append(sqrtConjX)


        h *= .5  
    
    #opening a workbook and adding the values of h, absError, relError into their own columns in xlsx file.
    wb = Workbook()
    sheet = wb.active
    r = 2
    for h in hList:
        sheet.cell(row= r, column=1).value = h
        r += 1
    r = 2
    for ap in apListS:
        sheet.cell(row=r, column= 2).value = ap
        r += 1
    r = 2
    for a in apListC:
        sheet.cell(row= r, column=3).value = a
        r += 1
    r = 2
    for  a in aListS:
        sheet.cell(row= r, column=4).value = a
        r += 1
    r = 2
    for a in aListC:
        sheet.cell(row=r, column= 5).value = a
        r += 1
    r = 2
    for re in rListS:
        sheet.cell(row= r, column=6).value = re
        r += 1
    r = 2
    for re in rListC:
        sheet.cell(row= r, column=7).value = re
        r += 1
    r = 2
    for re in rListC:
        sheet.cell(row= r, column=8).value = 7**(1/2)/14
        r += 1
    wb.save('TestApprox3.xlsx')

    

EvaluateApprox()
